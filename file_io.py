import pyexcel as pxl
from openpyxl import load_workbook
import pandas as pd
from pathlib import Path
import re

class FileIO:
    def read_file(path, no_strip=False):
        if (path.suffix == '.xls'):
            df = pd.read_excel(path)
            return df
            # pxl.save_as(file_name=path, dest_file_name=path.with_suffix('.xlsx'))
            # path = path.with_suffix('.xlsx')

        if no_strip:
            print(path)
            df = pd.read_excel(path, engine='openpyxl')
            return df

        workbook = load_workbook(path, data_only=False)
        sheet = workbook.active

        pattern = re.compile(r"^=\(?([0-9.]+)\)?$")
         
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    match = pattern.match(cell.value)
                    if match:
                        number = match.group(1)
                        try:
                             captured_value = float(number)
                        except ValueError:
                             captured_value = number

                        cell.value = captured_value

        workbook.save(path)

        df = pd.read_excel(path, engine='openpyxl')
        return df

    def write_file(df: pd.DataFrame, path: str):
        df.to_excel(path)
