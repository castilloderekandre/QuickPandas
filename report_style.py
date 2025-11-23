from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from pathlib import Path
from datetime import datetime

class ReportStyle:
  @classmethod
  def style_sheet(cls, path: Path):
    wb = load_workbook(path)
    active_sheet = wb.active

    if not active_sheet:
      return

    accounting_columns = [ 'H', 'I', 'J', 'M' ]
    columns_to_bold = ['B', 'C', 'D', 'F', 'J', 'P']
    sum_columns = [ 'H', 'J', 'M' ]
    color_fill = Color(theme=0, tint=-0.05, type='theme')
    apply_background = False
    for column_cells in active_sheet.iter_cols():
      apply_background = False

      for cell in column_cells:
        assert cell.row is not None
        active_sheet.row_dimensions[cell.row].height = 24
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        if apply_background:
          cell.fill = PatternFill(start_color=color_fill, end_color=color_fill, patternType='solid')
          apply_background = False
        else:
          apply_background = True


        if cell.column_letter in accounting_columns:
          cell.number_format = '_($* #,##0.00_);[Red]_($* (#,##0.00);_($* "-"??_);_(@_)'
        # if has_background_color:
          #  cell.fill = PatternFill(start_color=)

        if cell.column_letter in columns_to_bold:
          cell.font = Font(bold=True)
          cell.border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    max_row = active_sheet.max_row
    for sum_col in sum_columns:
      cell = active_sheet[sum_col + str(max_row + 1)]
      cell.value = f'=SUM({sum_col}2:{sum_col}{max_row})'
      cell.font = Font(bold=True)
      cell.number_format = '_($* #,##0.00_);[Red]_($* (#,##0.00);_($* "-"??_);_(@_)'
      cell.border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
      cell.alignment = Alignment(horizontal='center')

      if apply_background:
        cell.fill = PatternFill(start_color=color_fill, end_color=color_fill, patternType='solid')

    active_sheet.row_dimensions[max_row + 1].height = 24
    active_sheet.row_dimensions[max_row + 2].height = 24
    active_sheet.row_dimensions[max_row + 3].height = 24


    for column_cells in active_sheet.iter_cols():
      max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
      active_sheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    active_sheet.merge_cells(f'A{max_row+3}:P{max_row+3}')
    cell = active_sheet[f'A{max_row+3}']
    cell.value = f'Created On - {datetime.now().strftime("%m/%d/%Y")}'
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(size=20)

    wb.save(path)