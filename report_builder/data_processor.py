import re
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
pd.set_option('future.no_silent_downcasting', True)

class DataProcessor:
    inventory_fields = [
            'Product/Service Name', 
            'Sales Description', 
            'SKU',
            'Sales Price / Rate', 
            'Purchase Cost'
        ]

    @staticmethod
    def process_data(inventory: pd.DataFrame, expenses: pd.DataFrame):
        pass

    @staticmethod
    def parse_description(description, schema):
        schema_elements = []
        for identifier in schema:
            description = str(description)
            index = description.find(identifier) 

            if (index == -1):
                schema_elements.append('Not Found')
                continue

            

            end_index = description.find('\n', index)

            if (end_index == -1):
                end_index = len(description) - index

            index += len(identifier) + 1
            schema_elements.append(description[index:end_index])

        return schema_elements
    
    @staticmethod
    def parse_product_name(name):
        breakdown = []
        pattern = r":(\d+)\s+(\w+)\s+(.+)\(VIN#.+\)"
        capture_groups = re.search(pattern, name)
        if capture_groups:
            breakdown.extend(list(capture_groups.groups()))
        else:
            breakdown = ['-', '-', '-']
            
        return breakdown
    
    @staticmethod
    def parse_product(index, product: pd.Series, expenses: pd.DataFrame) -> list:
        schema: list[str] = [ 'Mileage' ]
        vin: str = product.SKU[-6:]
        in_inventory = product['Purchase Cost']
        if vin not in expenses.index:
            total_expenses = 0
        else:  
            total_expenses = expenses.loc[vin].sum()

        product_series = [
            vin, # LAST 6 OF VIN
            *DataProcessor.parse_product_name(product['Product/Service Name']), # YEAR, MAKE, MODEL
            *DataProcessor.parse_description(product['Sales Description'], schema), # MILEAGE (BASED ON SCHEMA)
            None, # LOCATION
            in_inventory, # in_inventory
            total_expenses, # EXPENSES
            in_inventory + total_expenses, # TOTAL INVESTED
            None, # MISC
            None, # SOURCED FROM
            product['Sales Price / Rate'], # SRP
            None, # DATE RECEIVED
            None, # OPEN INVOICE
            f'=TODAY() - N{index+2}', # AGE
        ]

        return product_series
    
    @staticmethod
    def clean_inventory(inventory: pd.DataFrame):
        return inventory.loc[inventory.Type == 'Inventory', DataProcessor.inventory_fields]

    @staticmethod
    def clean_expenses(expenses: pd.DataFrame):
        expenses = expenses[3:]
        expenses = expenses.T
        expenses.iloc[0, 0] = 'Class'
        expenses.columns = expenses.iloc[0]
        expenses = expenses[1:-1]
        expenses.set_index('Class', inplace=True)

        def clean_index(index):
            return str(index).split('.')[0]

        def clean_column(column):
            if column != str:
                column = str(column)
                return column.strip()

        expenses.index = expenses.index.map(clean_index)
        expenses.columns = expenses.columns.map(clean_column)
        
        expense_name_list: list[str] = []

        cost_of_goods_sold_reached: bool = False
        for series_name, _ in expenses.items():
            series_name = str(series_name)
            if not cost_of_goods_sold_reached:
                if series_name == 'Cost of goods sold':
                    cost_of_goods_sold_reached = True
                continue

            series_name_lower = str(series_name).lower()
            
            if series_name_lower == 'trucks purchased':
                continue

            if series_name_lower == 'total for cost of goods sold':
                break
            
            expense_name_list.append(str(series_name))
        
        existing_expense_name_list: list[str] = []
        for expense_name in expense_name_list:
            if expense_name in expenses.columns:
                existing_expense_name_list.append(expense_name)
        expenses = expenses[existing_expense_name_list]

        expenses = expenses.fillna(0).infer_objects(copy=False)

        return expenses

    @staticmethod
    def strip_formulas(path: Path):
        workbook = load_workbook(path, data_only=False)
        sheet = workbook.active
        if not sheet:
            return
        
        pattern = re.compile(r"^=\(?([0-9.]+)\)?$")
         
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    match = pattern.match(cell.value)
                    if match:
                        number = match.group(1)
                        try:
                             captured_value = float(number)
                        except ValueError:
                             captured_value = number

                        cell.value = captured_value

        workbook.save(path)
  